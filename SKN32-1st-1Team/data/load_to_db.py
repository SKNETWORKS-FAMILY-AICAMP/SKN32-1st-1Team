"""
전국 자동차 등록 현황 엑셀 -> MySQL DB 적재 스크립트
소나기 팀 | 2026

[사용법]
1. DB_CONFIG 수정 (password, db명)
2. DATA_DIR 에 YYYYMM_자동차등록통계.xlsx 추가
3. python load_to_db.py

[핵심 설계]
- load_history 테이블로 이미 적재된 파일 자동 스킵
- 새 파일만 골라서 적재 -> 매월 그냥 실행만 하면 됨
- ON DUPLICATE KEY UPDATE 로 중복 방지
- 시트19(연도별) 는 전체 파일 중 마지막에서만 1회 갱신
"""

import os
import glob
import re
import logging
from datetime import datetime

import pymysql
from openpyxl import load_workbook

from dotenv import load_dotenv

# ============================================================
# 설정
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'auto_data')

# .env 파일 로드
load_dotenv(os.path.join(BASE_DIR, '..', '.env'))

DB_CONFIG = {
    'host':       os.getenv('DB_HOST', 'localhost'),
    'port':       int(os.getenv('DB_PORT', 3306)),  # 포트는 정수형 변환이 필요합니다
    'user':       os.getenv('DB_USER', 'homework'),
    'password':   os.getenv('DB_PASSWORD'),
    'db':         os.getenv('DB_NAME'),
    'charset':    'utf8mb4',
    'autocommit': False,
}

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler('load_to_db.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

REGION_MAP = {
    '합계': '합계', '총계': '합계',
    '서울': '서울', '부산': '부산', '대구': '대구', '인천': '인천',
    '광주': '광주', '대전': '대전', '울산': '울산', '세종': '세종',
    '경기': '경기', '강원': '강원', '충북': '충북', '충남': '충남',
    '전북': '전북', '전남': '전남', '경북': '경북', '경남': '경남', '제주': '제주',
}


# ============================================================
# 유틸
# ============================================================
def safe_int(val) -> int:
    try:
        if val is None or str(val).strip() in ('', '-', 'None'):
            return 0
        return int(float(str(val).replace(',', '')))
    except Exception:
        return 0


def get_year_months(filepath: str) -> str:
    """
    파일명에서 YYYYMM 추출
    지원 형식:
      - 2025년_11월_자동차_등록자료_통계.xlsx  -> 202511
      - 202511_자동차등록통계.xlsx             -> 202511
    """
    base = os.path.basename(filepath)
    # 형식1: 2025년_11월_... -> YYYY년_MM월
    m = re.search(r'(\d{4})년_(\d{1,2})월', base)
    if m:
        return f"{m.group(1)}{int(m.group(2)):02d}"
    # 형식2: 202511_... -> 6자리 숫자
    m = re.search(r'(\d{6})', base)
    return m.group(1) if m else None


def get_files() -> list:
    files = sorted(glob.glob(os.path.join(DATA_DIR, '*자동차_등록자료_통계*.xlsx')))
    log.info(f'발견 파일: {len(files)}개')
    return files


# ============================================================
# 적재 이력 테이블
# ============================================================
def ensure_history_table(conn):
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS load_history (
                id           INT UNSIGNED NOT NULL AUTO_INCREMENT,
                year_months   CHAR(6)      NOT NULL COMMENT '적재 기준 년월 YYYYMM',
                filename     VARCHAR(200) NOT NULL,
                loaded_at    TIMESTAMP    NOT NULL DEFAULT CURRENT_TIMESTAMP,
                row_count    INT UNSIGNED NOT NULL DEFAULT 0,
                PRIMARY KEY (id),
                UNIQUE KEY uq_ym (year_months)
            ) ENGINE=InnoDB COMMENT='엑셀 파일 적재 이력'
        """)
    conn.commit()


def get_loaded_months(conn) -> set:
    with conn.cursor() as cur:
        cur.execute("SELECT year_months FROM load_history")
        return {row[0] for row in cur.fetchall()}


def mark_loaded(conn, year_months: str, filename: str, row_count: int):
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO load_history (year_months, filename, row_count)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE
                filename  = VALUES(filename),
                loaded_at = CURRENT_TIMESTAMP,
                row_count = VALUES(row_count)
        """, (year_months, filename, row_count))
    conn.commit()


# ============================================================
# 파서
# ============================================================
def parse_sheet01(ws, ym: str) -> list:
    rows = []
    CAR_TYPES = [
        ('승용',  2,  5), ('승합',  6,  9),
        ('화물', 10, 13), ('특수', 14, 17), ('합계', 18, 21),
    ]
    for row in ws.iter_rows(min_row=6, values_only=True):
        region = str(row[0]).strip() if row[0] else None
        if not region or region not in REGION_MAP:
            continue
        if region in ('월간증감', '년간증감'):
            break
        r = REGION_MAP[region]
        for ct, cs, ce in CAR_TYPES:
            cols = list(row[cs:ce + 1])
            if len(cols) < 4:
                continue
            rows.append({'year_months': ym, 'region': r, 'car_type': ct,
                         'cnt_official': safe_int(cols[0]), 'cnt_private': safe_int(cols[1]),
                         'cnt_business': safe_int(cols[2]), 'cnt_total': safe_int(cols[3])})
    return rows


def parse_sheet04(ws, ym: str) -> list:
    rows = []
    current_gender = None
    AGE_GROUPS = ['10대이하','20대','30대','40대','50대',
                  '60대','70대','80대','90대이상','법인및사업자','계']
    for row in ws.iter_rows(min_row=4, values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if row[1] else ''
        col0 = col0.replace(' ', '')
        col1 = col1.replace(' ', '')
        if col0 in ('남성', '여성', '기타'):
            current_gender = col0
            if col1 not in AGE_GROUPS:
                continue
            age_group = col1
        elif col0 in AGE_GROUPS:
            age_group = col0
        elif col1 in AGE_GROUPS:
            age_group = col1
        elif col0 == '합계':
            break
        else:
            continue
        if not current_gender:
            continue
        rows.append({'year_months': ym, 'region': '합계', 'gender': current_gender,
                     'age_group': age_group,
                     'registered_count': safe_int(row[2]) if len(row) > 2 else 0})
    return rows


def parse_sheet09(ws, ym: str) -> list:
    rows = []
    current_car = None
    for row in ws.iter_rows(min_row=5, values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if row[1] else ''
        if col0 in ('승용', '승합', '화물', '특수'):
            current_car = col0
            subtype = col1
        elif col0 == '총계':
            break
        elif col0 == '' and col1:
            subtype = col1
        elif col0 and col1:
            current_car = col0
            subtype = col1
        else:
            subtype = col0 if col0 else col1
        if not current_car or not subtype:
            continue
        rows.append({'year_months': ym, 'region': '합계', 'car_type': current_car,
                     'car_subtype': subtype,
                     'registered_count': safe_int(row[19]) if len(row) > 19 else 0})
    return rows


def parse_sheet10(ws, ym: str) -> list:
    rows = []
    current_fuel = None
    VALID_CAR = {'승용', '승합', '화물', '특수', '소계'}
    VALID_USG  = {'비사업용', '사업용', '계'}
    for row in ws.iter_rows(min_row=5, values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if row[1] else ''
        col2 = str(row[2]).strip() if row[2] else ''
        if col0 and col0 not in VALID_CAR:
            if col0 in ('총계', '합계'):
                current_fuel = None  # 합계/총계 섹션 시작 → 이후 행 무시
            else:
                current_fuel = col0
        if not current_fuel:
            continue
        if col1 not in VALID_CAR or col2 not in VALID_USG:
            continue
        rows.append({'year_months': ym, 'region': '합계', 'fuel_type': current_fuel,
                     'car_type': col1, 'usage_type': col2,
                     'registered_count': safe_int(row[20]) if len(row) > 20 else 0})
    return rows


def parse_sheet12(ws, ym: str) -> list:
    """배기량별(승용차) 등록현황 — 시트 12"""
    rows = []
    DISPLACEMENTS = [
        (2, '1000미만'),
        (3, '1000이상1600미만'),
        (4, '1600이상2000미만'),
        (5, '2000이상2500미만'),
        (6, '2500이상'),
        (7, '저속전기차'),
        (8, '전기차'),
    ]
    for row in ws.iter_rows(min_row=5, values_only=True):
        region = str(row[0]).strip() if row[0] else None
        if not region or region not in REGION_MAP:
            continue
        r = REGION_MAP[region]
        for col_idx, disp in DISPLACEMENTS:
            if col_idx < len(row):
                rows.append({'year_months': ym, 'region': r,
                             'displacement': disp,
                             'registered_count': safe_int(row[col_idx])})
    return rows


def parse_sheet15(ws, ym: str) -> list:
    """차령별·차종별·용도별 등록현황 — 시트 15"""
    rows = []
    CAR_TYPES = [
        ('승용',  1,  4), ('승합',  5,  8),
        ('화물',  9, 12), ('특수', 13, 16), ('합계', 17, 20),
    ]
    USAGE = ['관용', '자가용', '영업용', '계']
    for row in ws.iter_rows(min_row=5, values_only=True):
        try:
            model_year = int(row[0])
        except (TypeError, ValueError):
            continue
        if model_year < 1990 or model_year > 2030:
            continue
        for ct, cs, ce in CAR_TYPES:
            cols = list(row[cs:ce + 1])
            if len(cols) < 4:
                continue
            for i, u in enumerate(USAGE):
                rows.append({'year_months': ym, 'model_year': model_year,
                             'car_type': ct, 'usage_type': u,
                             'registered_count': safe_int(cols[i])})
    return rows


def parse_sheet19(ws) -> list:
    rows = []
    CAR_TYPES = [
        ('승용',  1,  4), ('승합',  5,  8),
        ('화물',  9, 12), ('특수', 13, 16), ('합계', 17, 20),
    ]
    USAGE = ['관용', '자가용', '영업용', '합계']
    for row in ws.iter_rows(min_row=5, values_only=True):
        try:
            year = int(row[0])
        except Exception:
            continue
        if year < 2007 or year > 2030:
            continue
        for ct, cs, ce in CAR_TYPES:
            cols = list(row[cs:ce + 1])
            if len(cols) < 4:
                continue
            for i, u in enumerate(USAGE):
                rows.append({'stat_year': year, 'car_type': ct,
                             'usage_type': u, 'registered_count': safe_int(cols[i])})
    return rows


def parse_sheet20(ws, ym: str) -> list:
    rows = []
    # col[1]은 빈 열 — 실제 데이터는 col[2]부터 시작
    CAR_TYPES = [
        ('승용',  2,  5), ('승합',  6,  9),
        ('화물', 10, 13), ('특수', 14, 17), ('합계', 18, 21),
    ]
    REG_TYPES = ['신조차', '수입차', '부활차', '계']
    for row in ws.iter_rows(min_row=5, values_only=True):
        region = str(row[0]).strip() if row[0] else None
        if not region or region not in REGION_MAP:
            continue
        r = REGION_MAP[region]
        for ct, cs, ce in CAR_TYPES:
            cols = list(row[cs:ce + 1])
            if len(cols) < 4:
                continue
            for i, rt in enumerate(REG_TYPES):
                rows.append({'year_months': ym, 'region': r, 'car_type': ct,
                             'reg_type': rt,
                             'registered_count': safe_int(cols[i])})
    return rows


# ============================================================
# DB INSERT
# ============================================================
def insert_many(conn, sql: str, data: list) -> int:
    if not data:
        return 0
    with conn.cursor() as cur:
        cur.executemany(sql, data)
    return len(data)


def upsert_region_stats(conn, rows):
    return insert_many(conn, """
        INSERT INTO car_region_stats
            (year_months,region,car_type,cnt_official,cnt_private,cnt_business,cnt_total)
        VALUES (%(year_months)s,%(region)s,%(car_type)s,
                %(cnt_official)s,%(cnt_private)s,%(cnt_business)s,%(cnt_total)s)
        ON DUPLICATE KEY UPDATE
            cnt_official=VALUES(cnt_official), cnt_private=VALUES(cnt_private),
            cnt_business=VALUES(cnt_business), cnt_total=VALUES(cnt_total)
    """, rows)


def upsert_gender_age(conn, rows):
    return insert_many(conn, """
        INSERT INTO car_gender_age_stats
            (year_months,region,gender,age_group,registered_count)
        VALUES (%(year_months)s,%(region)s,%(gender)s,%(age_group)s,%(registered_count)s)
        ON DUPLICATE KEY UPDATE registered_count=VALUES(registered_count)
    """, rows)


def upsert_type_detail(conn, rows):
    return insert_many(conn, """
        INSERT INTO car_type_detail_stats
            (year_months,region,car_type,car_subtype,registered_count)
        VALUES (%(year_months)s,%(region)s,%(car_type)s,%(car_subtype)s,%(registered_count)s)
        ON DUPLICATE KEY UPDATE registered_count=VALUES(registered_count)
    """, rows)


def upsert_fuel(conn, rows):
    return insert_many(conn, """
        INSERT INTO car_fuel_stats
            (year_months,region,fuel_type,car_type,usage_type,registered_count)
        VALUES (%(year_months)s,%(region)s,%(fuel_type)s,%(car_type)s,
                %(usage_type)s,%(registered_count)s)
        ON DUPLICATE KEY UPDATE registered_count=VALUES(registered_count)
    """, rows)


def upsert_yearly(conn, rows):
    return insert_many(conn, """
        INSERT INTO car_yearly_stats (stat_year,car_type,usage_type,registered_count)
        VALUES (%(stat_year)s,%(car_type)s,%(usage_type)s,%(registered_count)s)
        ON DUPLICATE KEY UPDATE registered_count=VALUES(registered_count)
    """, rows)


def upsert_displacement(conn, rows):
    return insert_many(conn, """
        INSERT INTO car_displacement_stats
            (year_months, region, displacement, registered_count)
        VALUES (%(year_months)s, %(region)s, %(displacement)s, %(registered_count)s)
        ON DUPLICATE KEY UPDATE registered_count=VALUES(registered_count)
    """, rows)


def upsert_car_age(conn, rows):
    return insert_many(conn, """
        INSERT INTO car_age_stats
            (year_months, model_year, car_type, usage_type, registered_count)
        VALUES (%(year_months)s, %(model_year)s, %(car_type)s,
                %(usage_type)s, %(registered_count)s)
        ON DUPLICATE KEY UPDATE registered_count=VALUES(registered_count)
    """, rows)


def upsert_new_reg(conn, rows):
    return insert_many(conn, """
        INSERT INTO car_new_registration
            (year_months,region,car_type,reg_type,registered_count)
        VALUES (%(year_months)s,%(region)s,%(car_type)s,%(reg_type)s,%(registered_count)s)
        ON DUPLICATE KEY UPDATE registered_count=VALUES(registered_count)
    """, rows)


# ============================================================
# 파일 처리
# ============================================================
def process_file(conn, filepath: str) -> int:
    ym = get_year_months(filepath)
    if not ym:
        log.warning(f'년월 추출 실패: {filepath}')
        return 0

    log.info(f'처리 시작: {os.path.basename(filepath)} ({ym})')
    total = 0

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        log.error(f'파일 열기 실패: {e}')
        return 0

    try:
        r01 = parse_sheet01(wb['01.통계표'], ym)
        total += upsert_region_stats(conn, r01)
        log.info(f'  [01] 시도별      : {len(r01):>5}행')

        r04 = parse_sheet04(wb['04.성별_연령별'], ym)
        total += upsert_gender_age(conn, r04)
        log.info(f'  [04] 성별·연령별 : {len(r04):>5}행')

        r09 = parse_sheet09(wb['09.차종별_유형별 현황'], ym)
        total += upsert_type_detail(conn, r09)
        log.info(f'  [09] 차종유형별  : {len(r09):>5}행')

        r10 = parse_sheet10(wb['10.연료별_등록현황'], ym)
        total += upsert_fuel(conn, r10)
        log.info(f'  [10] 연료별      : {len(r10):>5}행')

        r12 = parse_sheet12(wb['12.배기별(승용차)_등록현황'], ym)
        total += upsert_displacement(conn, r12)
        log.info(f'  [12] 배기량별    : {len(r12):>5}행')

        r15 = parse_sheet15(wb['15.차령별_차종별_용도별 등록현황'], ym)
        total += upsert_car_age(conn, r15)
        log.info(f'  [15] 차령별      : {len(r15):>5}행')

        r20 = parse_sheet20(wb['20.신규 등록현황(당월)'], ym)
        total += upsert_new_reg(conn, r20)
        log.info(f'  [20] 신규등록    : {len(r20):>5}행')

        conn.commit()
        log.info(f'  완료: 총 {total:,}행')

    except Exception as e:
        conn.rollback()
        log.error(f'오류 발생, 롤백: {e}', exc_info=True)
        total = 0
    finally:
        wb.close()

    return total


def update_yearly_stats(conn, latest_file: str):
    """연도별 장기 데이터를 전체 최신 파일에서 갱신 (항상 files[-1] 기준)"""
    log.info(f'연도별(장기) 갱신: {os.path.basename(latest_file)}')
    try:
        wb = load_workbook(latest_file, read_only=True, data_only=True)
        r19 = parse_sheet19(wb['19.연도별 자동차 등록현황'])
        upsert_yearly(conn, r19)
        conn.commit()
        log.info(f'  [19] 연도별(장기): {len(r19):>5}행')
        wb.close()
    except Exception as e:
        conn.rollback()
        log.error(f'연도별 갱신 오류: {e}', exc_info=True)


# ============================================================
# 메인
# ============================================================
def main():
    log.info('=' * 60)
    log.info('전국 자동차 등록 현황 DB 적재 시작')
    log.info(f'시작: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    log.info('=' * 60)

    files = get_files()
    if not files:
        log.error(f'파일 없음: {DATA_DIR}')
        return

    try:
        conn = pymysql.connect(**DB_CONFIG)
        log.info('DB 연결 성공')
    except Exception as e:
        log.error(f'DB 연결 실패: {e}')
        return

    try:
        # 콜레이션 통일 + UNSIGNED 차감 음수 허용 (SELECT 분석 쿼리와 동일 환경)
        with conn.cursor() as cur:
            cur.execute("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci")
            cur.execute("SET SESSION sql_mode = CONCAT(@@sql_mode, ',NO_UNSIGNED_SUBTRACTION')")
        conn.commit()

        ensure_history_table(conn)
        loaded_months = get_loaded_months(conn)

        # 신규 파일만 필터링 (이미 적재된 월은 스킵)
        new_files = [f for f in files if get_year_months(f) not in loaded_months]

        if not new_files:
            log.info('신규 파일 없음 — 모든 파일이 이미 적재되어 있습니다.')
            return

        log.info(f'신규 적재: {len(new_files)}개 '
                 f'(기적재 {len(files) - len(new_files)}개 스킵)')

        for filepath in new_files:
            ym = get_year_months(filepath)
            row_count = process_file(conn, filepath)
            if row_count > 0:
                mark_loaded(conn, ym, os.path.basename(filepath), row_count)

        # 연도별 데이터는 전체 파일 중 가장 최신 파일에서 1회 갱신
        # (중간 누락 파일 추가 시에도 항상 최신 스냅샷 기준으로 갱신)
        update_yearly_stats(conn, files[-1])

        # 최신 기준 날짜 출력
        with conn.cursor() as cur:
            cur.execute("SELECT MAX(year_months) FROM load_history")
            latest = cur.fetchone()[0]
        log.info(f'DB 최신 기준: {latest}')
        log.info('=' * 60)
        log.info(f'종료: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        log.info('=' * 60)

    finally:
        conn.close()


if __name__ == '__main__':
    main()